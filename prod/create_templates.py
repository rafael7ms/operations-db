"""Script to create all upload template Excel files."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ==================== SCHEDULE TEMPLATE ====================
def create_schedule_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule Upload'

    headers = [
        'Employee - ID', 'Date - Nominal Date', 'Earliest - Start',
        'Latest - Stop', 'Work - Code'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    sample_data = [
        [12345, '2024-01-15', '08:00', '17:00', 'REG'],
        [12346, '2024-01-15', '09:00', '18:00', 'REG'],
        [12347, '2024-01-15', '', '', 'OFF'],
        [12348, '2024-01-15', '22:00', '06:00', 'NIGHT'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:
                cell.number_format = 'YYYY-MM-DD'
            elif col_idx in [3, 4]:
                cell.number_format = 'HH:MM'

    instructions = [
        '1. Employee - ID: Employee numeric ID',
        '2. Date - Nominal Date: Schedule date (YYYY-MM-DD)',
        '3. Earliest - Start: Start time (HH:MM) or leave empty for OFF day',
        '4. Latest - Stop: Stop time (HH:MM)',
        '5. Work - Code: REG, NIGHT, OT, OFF, TRAIN, etc.',
    ]

    ws.column_dimensions['F'].width = 45
    ws.cell(row=1, column=6, value='Instructions')
    for i, inst in enumerate(instructions, 2):
        ws.cell(row=i, column=6, value=inst)

    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12

    wb.save('app/static/templates/schedule_upload_template.xlsx')
    print('Created: schedule_upload_template.xlsx')


# ==================== ATTENDANCE TEMPLATE ====================
def create_attendance_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Upload'

    headers = [
        'Employee - ID', 'Date', 'Check In', 'Check Out', 'Exception', 'Notes'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    sample_data = [
        [12345, '2024-01-15', '08:00', '17:00', '', 'Regular day'],
        [12346, '2024-01-15', '08:30', '17:30', 'OT', 'Overtime worked'],
        [12347, '2024-01-15', '08:00', '12:00', 'Sick', 'Sick leave'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 2:
                cell.number_format = 'YYYY-MM-DD'
            elif col_idx in [3, 4]:
                cell.number_format = 'HH:MM'

    instructions = [
        '1. Employee - ID: Employee numeric ID',
        '2. Date: Attendance date (YYYY-MM-DD)',
        '3. Check In: Check-in time (HH:MM)',
        '4. Check Out: Check-out time (HH:MM) - optional',
        '5. Exception: Type (OT, Sick, Vacation, etc.) - optional',
        '6. Notes: Additional notes - optional',
    ]

    ws.column_dimensions['G'].width = 45
    ws.cell(row=1, column=7, value='Instructions')
    for i, inst in enumerate(instructions, 2):
        ws.cell(row=i, column=7, value=inst)

    header_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

    wb.save('app/static/templates/attendance_upload_template.xlsx')
    print('Created: attendance_upload_template.xlsx')


# ==================== EMPLOYEE TEMPLATE ====================
def create_employee_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Upload'

    headers = [
        'Odoo ID', 'First Name', 'Last Name', 'Batch', 'Supervisor',
        'Manager', 'Shift', 'Department', 'Role', 'Hire Date',
        'Company Email', 'Tier', 'Agent ID', 'BO User', 'Axonify'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    sample_data = [
        [12345, 'John', 'Doe', '2024-01', 'Jane Smith', 'Bob Johnson',
         'Day', 'IT', 'Developer', '2024-01-15', 'john.doe@company.com', 2, 54321, 'bo_user1', 'axon123'],
        [12346, 'Jane', 'Roe', '2024-01', 'Jane Smith', 'Bob Johnson',
         'Day', 'IT', 'Analyst', '2024-01-16', 'jane.roe@company.com', 1, 54322, 'bo_user2', 'axon456'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 10:
                cell.number_format = 'YYYY-MM-DD'

    instructions = [
        '1. Odoo ID: Employee numeric ID (required)',
        '2-3. First/Last Name: Employee name (required)',
        '4. Batch: Batch number (required)',
        '5-6. Supervisor/Manager: Manager names (required)',
        '7. Shift: Day, Night, Rotating (required)',
        '8. Department: Department name (required)',
        '9. Role: Job role (required)',
        '10. Hire Date: Hire date (YYYY-MM-DD) (required)',
        '11. Company Email: Email address (required)',
        '12-15. Optional fields: Tier, Agent ID, BO User, Axonify',
    ]

    ws.column_dimensions['P'].width = 45
    ws.cell(row=1, column=16, value='Instructions')
    for i, inst in enumerate(instructions, 2):
        ws.cell(row=i, column=16, value=inst)

    header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['K'].width = 25

    wb.save('app/static/templates/employee_upload_template.xlsx')
    print('Created: employee_upload_template.xlsx')


# ==================== EXCEPTION TEMPLATE ====================
def create_exception_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Exception Upload'

    headers = [
        'Employee - ID', 'Exception Type', 'Start Date', 'End Date',
        'Work Code', 'Supervisor Override', 'Notes'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    sample_data = [
        [12345, 'Training', '2024-01-20', '2024-01-22', 'TRAIN', 'Jane Smith', 'New hire training'],
        [12346, 'Nesting', '2024-01-25', '2024-01-27', 'NEST', 'Jane Smith', 'Internal project'],
        [12347, 'Vacation', '2024-02-01', '2024-02-05', 'OFF', 'Bob Johnson', 'Personal time'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [3, 4]:
                cell.number_format = 'YYYY-MM-DD'

    instructions = [
        '1. Employee - ID: Employee numeric ID',
        '2. Exception Type: Training, Nesting, New Hire Training, Vacation, Sick, Personal, Bereavement, Work From Home',
        '3. Start Date: Start date (YYYY-MM-DD)',
        '4. End Date: End date (YYYY-MM-DD)',
        '5. Work Code: Work code during exception (optional)',
        '6. Supervisor Override: Supervisor name for approval (optional)',
        '7. Notes: Additional notes (optional)',
    ]

    ws.column_dimensions['H'].width = 45
    ws.cell(row=1, column=8, value='Instructions')
    for i, inst in enumerate(instructions, 2):
        ws.cell(row=i, column=8, value=inst)

    header_fill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20

    wb.save('app/static/templates/exception_upload_template.xlsx')
    print('Created: exception_upload_template.xlsx')


# ==================== REWARD TEMPLATE ====================
def create_reward_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reward Upload'

    headers = [
        'Employee - ID', 'Reason ID', 'Points', 'Date Awarded', 'Notes'
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    sample_data = [
        [12345, 1, 50, '2024-01-15', 'Excellent customer feedback'],
        [12346, 2, 25, '2024-01-16', 'Team player award'],
        [12347, 1, 100, '2024-01-17', 'Employee of the month'],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 4:
                cell.number_format = 'YYYY-MM-DD'

    instructions = [
        '1. Employee - ID: Employee numeric ID',
        '2. Reason ID: Reward reason ID from admin options',
        '3. Points: Number of points to award',
        '4. Date Awarded: Award date (YYYY-MM-DD)',
        '5. Notes: Additional notes (optional)',
    ]

    ws.column_dimensions['F'].width = 45
    ws.cell(row=1, column=6, value='Instructions')
    for i, inst in enumerate(instructions, 2):
        ws.cell(row=i, column=6, value=inst)

    header_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15

    wb.save('app/static/templates/reward_upload_template.xlsx')
    print('Created: reward_upload_template.xlsx')


if __name__ == '__main__':
    create_schedule_template()
    create_attendance_template()
    create_employee_template()
    create_exception_template()
    create_reward_template()
    print('\nAll templates created successfully!')
