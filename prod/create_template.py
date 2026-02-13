"""Script to create schedule upload template Excel file."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = 'Schedule Upload'

# Header row
headers = [
    'Employee - ID',      # Employee ID (required)
    'Date - Nominal Date', # Schedule date (required)
    'Earliest - Start',    # Start time (optional)
    'Latest - Stop',       # Stop time (optional)
    'Work - Code'          # Work code (optional)
]

for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Sample data rows
sample_data = [
    [12345, '2024-01-15', '08:00', '17:00', 'REG'],
    [12346, '2024-01-15', '09:00', '18:00', 'REG'],
    [12347, '2024-01-15', '', '', 'OFF'],  # Off day example
    [12348, '2024-01-15', '22:00', '06:00', 'NIGHT'],  # Overnight shift
    [12349, '2024-01-16', '08:00', '17:00', 'REG'],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 2:  # Date column
            cell.number_format = 'YYYY-MM-DD'
        elif col_idx in [3, 4]:  # Time columns
            cell.number_format = 'HH:MM'

# Add instructions in column F
ws.column_dimensions['F'].width = 45
ws.cell(row=1, column=6, value='Instructions')
ws.cell(row=2, column=6, value='1. Employee - ID: Employee numeric ID')
ws.cell(row=3, column=6, value='2. Date - Nominal Date: Schedule date (YYYY-MM-DD)')
ws.cell(row=4, column=6, value='3. Earliest - Start: Start time (HH:MM) or leave empty for OFF day')
ws.cell(row=5, column=6, value='4. Latest - Stop: Stop time (HH:MM)')
ws.cell(row=6, column=6, value='5. Work - Code: REG, NIGHT, OT, OFF, TRAIN, etc.')
ws.cell(row=7, column=6, value='6. Overnight shifts: If stop time < start time, stop date is next day')

# Format header row
header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

# Freeze first row
ws.freeze_panes = 'A2'

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 45

wb.save('app/static/templates/schedule_upload_template.xlsx')
print('Template created successfully!')
